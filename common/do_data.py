from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class DoData:
    @staticmethod
    def do_excel(file_path, sheet_name):
        workbook = load_workbook(file_path)
        ws: Worksheet = workbook[sheet_name]
        test_data = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=True):
            test_data.append(row)
        return test_data


if __name__ == '__main__':
    data = DoData.do_excel("../data/data.xlsx", "API")
    print(data)